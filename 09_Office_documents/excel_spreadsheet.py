import openpyxl
#
# workbook = openpyxl.load_workbook('book.xlsx')
# print(type(workbook))
# sheet = workbook['Sheet1']
# print(type(sheet))
# print(workbook.sheetnames)
#
# cell = sheet['B3']
# print(cell.value)
#
# print(sheet['E3'].value)
#
# print(sheet.cell(row=4, column=10))
#
# for i in range(4, 8):
#     print(1, sheet.cell(row=i, column=5).value)

# ==============================================

wb = openpyxl.Workbook()
print(wb)
print(wb.sheetnames)

sheet = wb['Sheet']
print(sheet['A1'].value)
sheet['A1'] = 42
sheet['A2'] = 'Hello'
sheet['A3'] = '30.10.2022'

wb.save('example.xlsx')
sheet2 = wb.create_sheet()
print(wb.sheetnames)

sheet2.title = "MyNewSheet"
print(wb.sheetnames)
wb.create_sheet(index=0, title='OtherSheet')
wb.save('example2.xlsx')
